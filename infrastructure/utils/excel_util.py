# Excel 文件的工具类
import os

from openpyxl import Workbook


# 导出excel文件
# path：文件导出路径
# file_name：文件名
# col_names: 列名
# row_names: 行名
# values：数据集, length(values) 等于 length(row_names), length(values[0]) 等于 length(col_names)
# 注意：openpyxl 的 sheet 下标从0开始，但是单元格的行、列是从1开始
# 示例:
#     列1  列2
# 行1  A   B
# 行2  C   D
def export(path, file_name, row_names, col_names, values):
    # 新建文件
    wb = Workbook()
    sheet = wb.create_sheet("sheet 1", 0)

    row_num = len(row_names)
    col_num = len(col_names)
    # 第1列,2~row_num+1 行标题
    for row_index in range(row_num):
        c = sheet.cell(row_index + 2, 1)
        c.value = row_names[row_index]
    # 第1行,2~col_num+1 是列标题
    for col_index in range(col_num):
        c = sheet.cell(1, col_index + 2)
        c.value = col_names[col_index]

    # 第2行\第2列开始,写入数据
    for row_index in range(row_num):
        for col_index in range(col_num):
            c = sheet.cell(row_index + 2, col_index + 2)
            c.value = values[row_index][col_index]
    # 保存文件
    file_path = path + "\\" + file_name + ".xlsx"
    wb.save(file_path)
    # 输出后打开
    os.startfile(file_path)
